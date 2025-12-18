# -*- coding: utf-8 -*-
"""
Report Service - Generates PDF and Excel reports for activity logs and messages
All timestamps are displayed in Argentina timezone (Buenos Aires)
"""

from io import BytesIO
from datetime import datetime
import json
import pytz

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from models import db, ActivityLog, Message


# Argentina timezone
BUENOS_AIRES_TZ = pytz.timezone('America/Argentina/Buenos_Aires')


def to_buenos_aires(dt):
    """Convert a datetime to Buenos Aires timezone"""
    if dt is None:
        return None
    if dt.tzinfo is None:
        # Assume UTC if no timezone
        dt = pytz.utc.localize(dt)
    return dt.astimezone(BUENOS_AIRES_TZ)


def format_datetime_ar(dt):
    """Format datetime for Argentina display"""
    if dt is None:
        return '-'
    dt_ar = to_buenos_aires(dt)
    return dt_ar.strftime('%d/%m/%Y %H:%M')


def log_activity(user_id, action_type, entity_type, entity_id=None, entity_name=None, details=None):
    """
    Log a user activity
    
    Args:
        user_id: ID of the user performing the action
        action_type: Type of action (create, update, delete)
        entity_type: Type of entity (broker, portfolio, investment, stock, message, portfolio_stock)
        entity_id: ID of the entity (optional)
        entity_name: Name/description of the entity (optional)
        details: Additional details as dict (will be stored as JSON)
    """
    try:
        activity = ActivityLog(
            user_id=user_id,
            action_type=action_type,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_name=entity_name,
            details=json.dumps(details) if details else None
        )
        db.session.add(activity)
        db.session.commit()
        return activity
    except Exception as e:
        db.session.rollback()
        print(f"Error logging activity: {e}")
        return None


def get_activities(start_date=None, end_date=None):
    """
    Get activity logs filtered by date range
    
    Args:
        start_date: Start date (datetime)
        end_date: End date (datetime)
    
    Returns:
        List of ActivityLog objects
    """
    query = ActivityLog.query.order_by(ActivityLog.created_at.desc())
    
    if start_date:
        query = query.filter(ActivityLog.created_at >= start_date)
    if end_date:
        # Include the entire end day
        end_of_day = datetime.combine(end_date, datetime.max.time())
        query = query.filter(ActivityLog.created_at <= end_of_day)
    
    return query.all()


def get_messages(start_date=None, end_date=None):
    """
    Get messages filtered by date range
    
    Args:
        start_date: Start date (datetime)
        end_date: End date (datetime)
    
    Returns:
        List of Message objects
    """
    query = Message.query.order_by(Message.created_at.desc())
    
    if start_date:
        query = query.filter(Message.created_at >= start_date)
    if end_date:
        end_of_day = datetime.combine(end_date, datetime.max.time())
        query = query.filter(Message.created_at <= end_of_day)
    
    return query.all()


# ==================== PDF GENERATION ====================

def _get_action_text(action_type):
    """Translate action type to Spanish"""
    translations = {
        'create': 'Creación',
        'update': 'Actualización',
        'delete': 'Eliminación'
    }
    return translations.get(action_type, action_type)


def _get_entity_text(entity_type):
    """Translate entity type to Spanish"""
    translations = {
        'broker': 'Broker',
        'portfolio': 'Cartera',
        'investment': 'Inversión',
        'stock': 'Activo',
        'portfolio_stock': 'Activo en Cartera',
        'message': 'Mensaje'
    }
    return translations.get(entity_type, entity_type)


def generate_activities_pdf(activities, start_date=None, end_date=None):
    """
    Generate a professional PDF report of activities
    
    Returns:
        BytesIO buffer containing the PDF
    """
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, KeepTogether
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus.frames import Frame
    from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
    from functools import partial
    
    buffer = BytesIO()
    
    # Custom colors - Professional dark blue theme
    PRIMARY_COLOR = colors.HexColor('#1a365d')  # Dark navy blue
    SECONDARY_COLOR = colors.HexColor('#2c5282')  # Medium blue
    ACCENT_COLOR = colors.HexColor('#3182ce')  # Bright blue
    HEADER_BG = colors.HexColor('#1a365d')
    ROW_ALT = colors.HexColor('#f7fafc')  # Very light gray
    BORDER_COLOR = colors.HexColor('#e2e8f0')  # Light border
    TEXT_DARK = colors.HexColor('#1a202c')
    TEXT_MUTED = colors.HexColor('#718096')
    
    def add_page_header_footer(canvas, doc, start_date, end_date):
        """Add header and footer to each page"""
        canvas.saveState()
        width, height = landscape(A4)
        
        # ===== HEADER =====
        # Header background
        canvas.setFillColor(PRIMARY_COLOR)
        canvas.rect(0, height - 60, width, 60, fill=True, stroke=False)
        
        # Header accent line
        canvas.setFillColor(ACCENT_COLOR)
        canvas.rect(0, height - 64, width, 4, fill=True, stroke=False)
        
        # Company/App name
        canvas.setFillColor(colors.white)
        canvas.setFont('Helvetica-Bold', 20)
        canvas.drawString(30, height - 40, "Sistema de Inversiones")
        
        # Report type badge
        canvas.setFont('Helvetica', 11)
        canvas.drawString(30, height - 55, "Reporte de Movimientos")
        
        # Date range on right
        canvas.setFont('Helvetica', 10)
        date_text = ""
        if start_date and end_date:
            date_text = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        elif start_date:
            date_text = f"Desde {start_date.strftime('%d/%m/%Y')}"
        elif end_date:
            date_text = f"Hasta {end_date.strftime('%d/%m/%Y')}"
        else:
            date_text = "Todos los registros"
        canvas.drawRightString(width - 30, height - 35, f"Período: {date_text}")
        
        # Generation timestamp
        now_ar = to_buenos_aires(datetime.utcnow())
        canvas.setFont('Helvetica', 9)
        canvas.drawRightString(width - 30, height - 50, f"Generado: {now_ar.strftime('%d/%m/%Y %H:%M')} hs")
        
        # ===== FOOTER =====
        # Footer line
        canvas.setStrokeColor(BORDER_COLOR)
        canvas.setLineWidth(1)
        canvas.line(30, 35, width - 30, 35)
        
        # Page number
        canvas.setFillColor(TEXT_MUTED)
        canvas.setFont('Helvetica', 9)
        page_num = canvas.getPageNumber()
        canvas.drawCentredString(width / 2, 20, f"Página {page_num}")
        
        # Footer text left
        canvas.drawString(30, 20, "Sistema de Inversiones - Reporte Confidencial")
        
        # Footer text right
        canvas.drawRightString(width - 30, 20, now_ar.strftime('%d/%m/%Y'))
        
        canvas.restoreState()
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=80,  # Space for header
        bottomMargin=50  # Space for footer
    )
    
    styles = getSampleStyleSheet()
    elements = []
    
    # Summary stats card
    stats_style = ParagraphStyle(
        'Stats',
        parent=styles['Normal'],
        fontSize=11,
        textColor=TEXT_DARK,
        spaceAfter=5
    )
    
    # Count activities by type
    action_counts = {}
    entity_counts = {}
    for activity in activities:
        action = _get_action_text(activity.action_type)
        entity = _get_entity_text(activity.entity_type)
        action_counts[action] = action_counts.get(action, 0) + 1
        entity_counts[entity] = entity_counts.get(entity, 0) + 1
    
    # Stats summary box
    if activities:
        stats_data = [
            [
                Paragraph(f"<b>Total de Movimientos:</b> {len(activities)}", stats_style),
                Paragraph(f"<b>Acciones:</b> {', '.join(f'{k} ({v})' for k, v in action_counts.items())}", stats_style),
            ]
        ]
        stats_table = Table(stats_data, colWidths=[12*cm, 14*cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#edf2f7')),
            ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 15),
            ('RIGHTPADDING', (0, 0), (-1, -1), 15),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 15))
    
    if not activities:
        empty_style = ParagraphStyle(
            'Empty',
            parent=styles['Normal'],
            fontSize=12,
            textColor=TEXT_MUTED,
            alignment=TA_CENTER,
            spaceBefore=50,
            spaceAfter=50
        )
        elements.append(Paragraph("No hay movimientos en el período seleccionado.", empty_style))
    else:
        # Table header with icons/symbols
        header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.white,
            fontName='Helvetica-Bold'
        )
        
        data = [[
            Paragraph('FECHA/HORA', header_style),
            Paragraph('USUARIO', header_style),
            Paragraph('ACCIÓN', header_style),
            Paragraph('ENTIDAD', header_style),
            Paragraph('NOMBRE', header_style),
            Paragraph('DETALLES', header_style)
        ]]
        
        # Cell style for data
        cell_style = ParagraphStyle(
            'Cell',
            parent=styles['Normal'],
            fontSize=8,
            textColor=TEXT_DARK,
            leading=11
        )
        
        # Table rows
        for activity in activities:
            username = activity.user.username if activity.user else 'N/A'
            details_text = ''
            if activity.details:
                try:
                    details_dict = json.loads(activity.details)
                    # Format details in a human-readable way
                    formatted_parts = []
                    for k, v in details_dict.items():
                        # Translate common field names to Spanish
                        field_translations = {
                            'quantity': 'Cantidad',
                            'price': 'Precio',
                            'type': 'Tipo',
                            'amount': 'Monto',
                            'broker_id': 'Broker ID',
                            'investment_id': 'Inversión ID',
                            'portfolio_id': 'Cartera ID',
                            # New Spanish field names
                            'broker': 'Broker',
                            'inversión': 'Inversión',
                            'cartera': 'Cartera'
                        }
                        field_name = field_translations.get(k, k.replace('_', ' ').title())
                        
                        # Format numbers nicely
                        if isinstance(v, (int, float)):
                            if k in ['price', 'amount']:
                                formatted_value = f"${v:,.2f}"
                            elif k == 'quantity':
                                formatted_value = f"{v:,.0f}" if v == int(v) else f"{v:,.2f}"
                            else:
                                formatted_value = str(v)
                        else:
                            formatted_value = str(v)
                        
                        formatted_parts.append(f"{field_name}: {formatted_value}")
                    details_text = ' | '.join(formatted_parts)
                except:
                    details_text = activity.details[:50]
            
            # Action badge color
            action_text = _get_action_text(activity.action_type)
            if activity.action_type == 'create':
                action_display = f"✓ {action_text}"
            elif activity.action_type == 'update':
                action_display = f"✎ {action_text}"
            elif activity.action_type == 'delete':
                action_display = f"✗ {action_text}"
            else:
                action_display = action_text
            
            data.append([
                Paragraph(format_datetime_ar(activity.created_at), cell_style),
                Paragraph(username, cell_style),
                Paragraph(action_display, cell_style),
                Paragraph(_get_entity_text(activity.entity_type), cell_style),
                Paragraph(activity.entity_name or '-', cell_style),
                Paragraph(details_text[:60] if details_text else '-', cell_style)
            ])
        
        # Create table with professional styling
        col_widths = [3.2*cm, 2.8*cm, 2.5*cm, 2.8*cm, 5.5*cm, 7*cm]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Professional table style
        table_style = [
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            
            # Borders - clean minimal style
            ('LINEBELOW', (0, 0), (-1, 0), 2, ACCENT_COLOR),
            ('LINEBELOW', (0, 1), (-1, -2), 0.5, BORDER_COLOR),
            ('LINEBELOW', (0, -1), (-1, -1), 1, BORDER_COLOR),
            ('LINEBEFORE', (0, 0), (0, -1), 0.5, BORDER_COLOR),
            ('LINEAFTER', (-1, 0), (-1, -1), 0.5, BORDER_COLOR),
            
            # Alignment
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        
        # Alternating row colors
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.append(('BACKGROUND', (0, i), (-1, i), ROW_ALT))
            else:
                table_style.append(('BACKGROUND', (0, i), (-1, i), colors.white))
        
        table.setStyle(TableStyle(table_style))
        elements.append(table)
    
    # Build with custom header/footer
    doc.build(
        elements,
        onFirstPage=partial(add_page_header_footer, start_date=start_date, end_date=end_date),
        onLaterPages=partial(add_page_header_footer, start_date=start_date, end_date=end_date)
    )
    buffer.seek(0)
    return buffer


def generate_messages_pdf(messages, start_date=None, end_date=None):
    """
    Generate a professional PDF report of messages
    
    Returns:
        BytesIO buffer containing the PDF
    """
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, KeepTogether
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from functools import partial
    
    buffer = BytesIO()
    
    # Custom colors - Professional purple theme for messages
    PRIMARY_COLOR = colors.HexColor('#44337a')  # Dark purple
    SECONDARY_COLOR = colors.HexColor('#553c9a')  # Medium purple
    ACCENT_COLOR = colors.HexColor('#805ad5')  # Bright purple
    HEADER_BG = colors.HexColor('#44337a')
    ROW_ALT = colors.HexColor('#faf5ff')  # Very light purple
    BORDER_COLOR = colors.HexColor('#e9d8fd')  # Light purple border
    TEXT_DARK = colors.HexColor('#1a202c')
    TEXT_MUTED = colors.HexColor('#718096')
    
    def add_page_header_footer(canvas, doc, start_date, end_date):
        """Add header and footer to each page"""
        canvas.saveState()
        width, height = landscape(A4)
        
        # ===== HEADER =====
        # Header background
        canvas.setFillColor(PRIMARY_COLOR)
        canvas.rect(0, height - 60, width, 60, fill=True, stroke=False)
        
        # Header accent line
        canvas.setFillColor(ACCENT_COLOR)
        canvas.rect(0, height - 64, width, 4, fill=True, stroke=False)
        
        # Company/App name
        canvas.setFillColor(colors.white)
        canvas.setFont('Helvetica-Bold', 20)
        canvas.drawString(30, height - 40, "Sistema de Inversiones")
        
        # Report type badge
        canvas.setFont('Helvetica', 11)
        canvas.drawString(30, height - 55, "Reporte de Mensajes")
        
        # Date range on right
        canvas.setFont('Helvetica', 10)
        date_text = ""
        if start_date and end_date:
            date_text = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        elif start_date:
            date_text = f"Desde {start_date.strftime('%d/%m/%Y')}"
        elif end_date:
            date_text = f"Hasta {end_date.strftime('%d/%m/%Y')}"
        else:
            date_text = "Todos los registros"
        canvas.drawRightString(width - 30, height - 35, f"Período: {date_text}")
        
        # Generation timestamp
        now_ar = to_buenos_aires(datetime.utcnow())
        canvas.setFont('Helvetica', 9)
        canvas.drawRightString(width - 30, height - 50, f"Generado: {now_ar.strftime('%d/%m/%Y %H:%M')} hs")
        
        # ===== FOOTER =====
        # Footer line
        canvas.setStrokeColor(BORDER_COLOR)
        canvas.setLineWidth(1)
        canvas.line(30, 35, width - 30, 35)
        
        # Page number
        canvas.setFillColor(TEXT_MUTED)
        canvas.setFont('Helvetica', 9)
        page_num = canvas.getPageNumber()
        canvas.drawCentredString(width / 2, 20, f"Página {page_num}")
        
        # Footer text left
        canvas.drawString(30, 20, "Sistema de Inversiones - Reporte Confidencial")
        
        # Footer text right
        canvas.drawRightString(width - 30, 20, now_ar.strftime('%d/%m/%Y'))
        
        canvas.restoreState()
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=80,  # Space for header
        bottomMargin=50  # Space for footer
    )
    
    styles = getSampleStyleSheet()
    elements = []
    
    # Summary stats card
    stats_style = ParagraphStyle(
        'Stats',
        parent=styles['Normal'],
        fontSize=11,
        textColor=TEXT_DARK,
        spaceAfter=5
    )
    
    # Count messages by type
    type_counts = {}
    for msg in messages:
        msg_type = msg.message_type or 'general'
        type_counts[msg_type] = type_counts.get(msg_type, 0) + 1
    
    # Stats summary box
    if messages:
        type_summary = ', '.join(f'{k.title()} ({v})' for k, v in type_counts.items())
        stats_data = [
            [
                Paragraph(f"<b>Total de Mensajes:</b> {len(messages)}", stats_style),
                Paragraph(f"<b>Por tipo:</b> {type_summary}", stats_style),
            ]
        ]
        stats_table = Table(stats_data, colWidths=[10*cm, 16*cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f3e8ff')),
            ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 15),
            ('RIGHTPADDING', (0, 0), (-1, -1), 15),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 15))
    
    if not messages:
        empty_style = ParagraphStyle(
            'Empty',
            parent=styles['Normal'],
            fontSize=12,
            textColor=TEXT_MUTED,
            alignment=TA_CENTER,
            spaceBefore=50,
            spaceAfter=50
        )
        elements.append(Paragraph("No hay mensajes en el período seleccionado.", empty_style))
    else:
        # Table header
        header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.white,
            fontName='Helvetica-Bold'
        )
        
        data = [[
            Paragraph('FECHA/HORA', header_style),
            Paragraph('AUTOR', header_style),
            Paragraph('TIPO', header_style),
            Paragraph('ENTIDAD ASOCIADA', header_style),
            Paragraph('CONTENIDO', header_style)
        ]]
        
        # Cell style for data
        cell_style = ParagraphStyle(
            'Cell',
            parent=styles['Normal'],
            fontSize=8,
            textColor=TEXT_DARK,
            leading=11
        )
        
        # Table rows
        for msg in messages:
            author = msg.author.full_name or msg.author.username if msg.author else 'N/A'
            
            # Determine associated entity with icon
            entity = '-'
            if msg.broker_id and msg.broker:
                entity = f"🏢 {msg.broker.name}"
            elif msg.investment_id and msg.investment:
                entity = f"💰 {msg.investment.name}"
            elif msg.portfolio_id and msg.portfolio:
                entity = f"📊 {msg.portfolio.name}"
            
            # Type with icon
            msg_type = msg.message_type or 'general'
            type_icons = {
                'broker': '🏢',
                'investment': '💰',
                'portfolio': '📊',
                'general': '💬'
            }
            type_display = f"{type_icons.get(msg_type, '💬')} {msg_type.title()}"
            
            # Truncate content for table
            content = msg.content[:120] + '...' if len(msg.content) > 120 else msg.content
            
            data.append([
                Paragraph(format_datetime_ar(msg.created_at), cell_style),
                Paragraph(author, cell_style),
                Paragraph(type_display, cell_style),
                Paragraph(entity, cell_style),
                Paragraph(content, cell_style)
            ])
        
        # Create table with professional styling
        col_widths = [3.2*cm, 3*cm, 2.5*cm, 5*cm, 10*cm]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Professional table style
        table_style = [
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            
            # Borders - clean minimal style
            ('LINEBELOW', (0, 0), (-1, 0), 2, ACCENT_COLOR),
            ('LINEBELOW', (0, 1), (-1, -2), 0.5, BORDER_COLOR),
            ('LINEBELOW', (0, -1), (-1, -1), 1, BORDER_COLOR),
            ('LINEBEFORE', (0, 0), (0, -1), 0.5, BORDER_COLOR),
            ('LINEAFTER', (-1, 0), (-1, -1), 0.5, BORDER_COLOR),
            
            # Alignment
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        
        # Alternating row colors
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.append(('BACKGROUND', (0, i), (-1, i), ROW_ALT))
            else:
                table_style.append(('BACKGROUND', (0, i), (-1, i), colors.white))
        
        table.setStyle(TableStyle(table_style))
        elements.append(table)
    
    # Build with custom header/footer
    doc.build(
        elements,
        onFirstPage=partial(add_page_header_footer, start_date=start_date, end_date=end_date),
        onLaterPages=partial(add_page_header_footer, start_date=start_date, end_date=end_date)
    )
    buffer.seek(0)
    return buffer


# ==================== EXCEL GENERATION ====================

def _style_excel_header(ws, header_row=1):
    """Apply styling to Excel header row"""
    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def generate_activities_excel(activities, start_date=None, end_date=None):
    """
    Generate an Excel report of activities
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    
    # Header
    headers = ['Fecha/Hora', 'Usuario', 'Acción', 'Entidad', 'Nombre', 'ID Entidad', 'Detalles']
    ws.append(headers)
    _style_excel_header(ws)
    
    # Data rows
    for activity in activities:
        username = activity.user.username if activity.user else 'N/A'
        details_text = ''
        if activity.details:
            try:
                details_dict = json.loads(activity.details)
                details_text = ', '.join(f"{k}: {v}" for k, v in details_dict.items())
            except:
                details_text = activity.details
        
        ws.append([
            format_datetime_ar(activity.created_at),
            username,
            _get_action_text(activity.action_type),
            _get_entity_text(activity.entity_type),
            activity.entity_name or '-',
            activity.entity_id or '-',
            details_text or '-'
        ])
    
    # Auto-adjust column widths
    for col_num, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add metadata sheet
    ws_meta = wb.create_sheet(title="Info")
    ws_meta.append(["Reporte de Movimientos"])
    ws_meta.append([])
    
    date_range_text = "Período: "
    if start_date and end_date:
        date_range_text += f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    elif start_date:
        date_range_text += f"Desde {start_date.strftime('%d/%m/%Y')}"
    elif end_date:
        date_range_text += f"Hasta {end_date.strftime('%d/%m/%Y')}"
    else:
        date_range_text += "Todos los registros"
    ws_meta.append([date_range_text])
    
    now_ar = to_buenos_aires(datetime.utcnow())
    ws_meta.append([f"Generado: {now_ar.strftime('%d/%m/%Y %H:%M')} (hora Argentina)"])
    ws_meta.append([f"Total de movimientos: {len(activities)}"])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_messages_excel(messages, start_date=None, end_date=None):
    """
    Generate an Excel report of messages
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mensajes"
    
    # Header
    headers = ['Fecha/Hora', 'Autor', 'Tipo', 'Broker', 'Inversión', 'Cartera', 'Contenido']
    ws.append(headers)
    
    # Style header with purple for messages
    header_fill = PatternFill(start_color='9b59b6', end_color='9b59b6', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for msg in messages:
        author = msg.author.full_name or msg.author.username if msg.author else 'N/A'
        
        broker_name = msg.broker.name if msg.broker_id and msg.broker else '-'
        investment_name = msg.investment.name if msg.investment_id and msg.investment else '-'
        portfolio_name = msg.portfolio.name if msg.portfolio_id and msg.portfolio else '-'
        
        ws.append([
            format_datetime_ar(msg.created_at),
            author,
            msg.message_type or 'general',
            broker_name,
            investment_name,
            portfolio_name,
            msg.content
        ])
    
    # Auto-adjust column widths
    for col_num, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add metadata sheet
    ws_meta = wb.create_sheet(title="Info")
    ws_meta.append(["Reporte de Mensajes"])
    ws_meta.append([])
    
    date_range_text = "Período: "
    if start_date and end_date:
        date_range_text += f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    elif start_date:
        date_range_text += f"Desde {start_date.strftime('%d/%m/%Y')}"
    elif end_date:
        date_range_text += f"Hasta {end_date.strftime('%d/%m/%Y')}"
    else:
        date_range_text += "Todos los registros"
    ws_meta.append([date_range_text])
    
    now_ar = to_buenos_aires(datetime.utcnow())
    ws_meta.append([f"Generado: {now_ar.strftime('%d/%m/%Y %H:%M')} (hora Argentina)"])
    ws_meta.append([f"Total de mensajes: {len(messages)}"])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
